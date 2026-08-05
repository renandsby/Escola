"""
Gerador de PDFs para relatórios escolares
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import io
import qrcode


def generate_student_report_pdf(student, grades, attendance):
    """Gera PDF do boletim do aluno"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    elements = []

    # Cabeçalho
    elements.append(Paragraph("BOLETIM ESCOLAR", title_style))
    elements.append(Spacer(1, 0.3*inch))

    # Dados do aluno
    info_data = [
        ['Nome:', f"{student.user.get_full_name()}"],
        ['Matrícula:', student.registration_number],
        ['Data de Nascimento:', str(student.birth_date)],
        ['Gênero:', student.gender],
    ]

    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4b5563')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    # Tabela de notas
    if grades:
        elements.append(Paragraph("DESEMPENHO POR DISCIPLINA", ParagraphStyle(
            'SubHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        )))
        elements.append(Spacer(1, 0.2*inch))

        grades_data = [['Disciplina', '1º Per.', '2º Per.', '3º Per.', '4º Per.', 'Média', 'Status']]

        for grade in grades:
            avg = f"{grade.get_average():.1f}" if grade.get_average() else "-"
            grades_data.append([
                grade.subject.name,
                f"{grade.first_period}" if grade.first_period else "-",
                f"{grade.second_period}" if grade.second_period else "-",
                f"{grade.third_period}" if grade.third_period else "-",
                f"{grade.fourth_period}" if grade.fourth_period else "-",
                avg,
                grade.get_status_display().title()
            ])

        grades_table = Table(grades_data, colWidths=[2.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
        grades_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(grades_table)

    elements.append(Spacer(1, 0.3*inch))

    # Frequência
    if attendance:
        elements.append(Paragraph("RESUMO DE FREQUÊNCIA", ParagraphStyle(
            'SubHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        )))
        elements.append(Spacer(1, 0.2*inch))

        present = sum(1 for a in attendance if a.status == 'present')
        absent = sum(1 for a in attendance if a.status == 'absent')
        justified = sum(1 for a in attendance if a.status == 'justified')
        total = len(attendance)
        percent = (present / total * 100) if total > 0 else 0

        freq_data = [
            ['Total de Aulas', 'Presentes', 'Ausentes', 'Justificados', 'Frequência %'],
            [str(total), str(present), str(absent), str(justified), f"{percent:.1f}%"]
        ]

        freq_table = Table(freq_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        freq_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f0fdf4')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(freq_table)

    elements.append(Spacer(1, 0.5*inch))

    # Rodapé
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#9ca3af'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Documento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_student_card_pdf(student):
    """Gera PDF da carteirinha do aluno com QR Code"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(4*inch, 6*inch), topMargin=0.25*inch, bottomMargin=0.25*inch)

    styles = getSampleStyleSheet()
    elements = []

    # Gerar QR Code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(f"MAT:{student.registration_number}|{student.user.get_full_name()}")
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)

    # Cabeçalho colorido
    header_data = [['CARTEIRINHA ESCOLAR']]
    header_table = Table(header_data, colWidths=[3.5*inch])
    header_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica-Bold', 14),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#3b82f6')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.2*inch))

    # Informações
    info_text = f"""
    <b>Nome:</b> {student.user.get_full_name()}<br/>
    <b>Matrícula:</b> {student.registration_number}<br/>
    <b>Data de Nascimento:</b> {student.birth_date.strftime('%d/%m/%Y')}<br/>
    """

    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))

    # QR Code
    from reportlab.platypus import Image as RLImage
    qr_img_rl = RLImage(qr_buffer, width=1.5*inch, height=1.5*inch)
    elements.append(qr_img_rl)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(students, grades):
    """Gera relatório em Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boletim"

        # Cabeçalho
        ws['A1'] = "RELATÓRIO DE NOTAS - BOLETIM CONSOLIDADO"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')

        # Colunas
        headers = ['Nome', 'Matrícula', 'Disciplina', '1º Per.', '2º Per.', '3º Per.', '4º Per.', 'Média']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dados
        row = 4
        for grade in grades:
            ws.cell(row=row, column=1).value = grade.student.user.get_full_name()
            ws.cell(row=row, column=2).value = grade.student.registration_number
            ws.cell(row=row, column=3).value = grade.subject.name
            ws.cell(row=row, column=4).value = grade.first_period
            ws.cell(row=row, column=5).value = grade.second_period
            ws.cell(row=row, column=6).value = grade.third_period
            ws.cell(row=row, column=7).value = grade.fourth_period
            ws.cell(row=row, column=8).value = grade.get_average()
            row += 1

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except ImportError:
        return None


def generate_csv_report(students, grades):
    """Gera relatório em CSV"""
    import csv

    buffer = io.BytesIO()
    text_buffer = io.StringIO()

    writer = csv.writer(text_buffer)
    writer.writerow(['Nome', 'Matrícula', 'Disciplina', '1º Período', '2º Período', '3º Período', '4º Período', 'Média'])

    for grade in grades:
        writer.writerow([
            grade.student.user.get_full_name(),
            grade.student.registration_number,
            grade.subject.name,
            grade.first_period,
            grade.second_period,
            grade.third_period,
            grade.fourth_period,
            grade.get_average()
        ])

    buffer.write(text_buffer.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer
