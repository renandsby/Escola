"""
Gerador de PDFs para relatórios escolares
"""

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from datetime import datetime
import io
import qrcode

from apps.class_diary.models import AttendanceStatus
from core.validators import format_cpf


def _student_display_name(student):
    return student.full_name or (
        student.user.get_full_name() if student.user_id else student.unique_municipal_id
    )


def generate_student_report_pdf(student, grades, attendance):
    """Gera PDF do boletim do aluno"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    elements = []

    elements.append(Paragraph("BOLETIM ESCOLAR", title_style))
    elements.append(Spacer(1, 0.3*inch))

    info_data = [
        ['Nome:', _student_display_name(student)],
        ['CPF:', format_cpf(student.cpf) if student.cpf else 'Não informado'],
        ['ID Municipal:', student.unique_municipal_id],
        ['Data de Nascimento:', str(student.birth_date)],
        ['Gênero:', student.gender or '-'],
    ]

    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4b5563')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    if grades:
        elements.append(Paragraph("DESEMPENHO POR DISCIPLINA", ParagraphStyle(
            'SubHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        )))
        elements.append(Spacer(1, 0.2*inch))

        grades_data = [['Disciplina', 'Período', 'Nota', 'Recuperação', 'Final', 'Efetiva']]

        for grade in grades:
            effective = grade.get_effective_score()
            grades_data.append([
                grade.subject.name,
                grade.academic_period.name,
                f"{grade.score}",
                f"{grade.recovery_score}" if grade.recovery_score is not None else "-",
                f"{grade.final_score}" if grade.final_score is not None else "-",
                f"{effective}" if effective is not None else "-",
            ])

        grades_table = Table(
            grades_data,
            colWidths=[2*inch, 1.2*inch, 0.8*inch, 1*inch, 0.8*inch, 0.8*inch],
        )
        grades_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(grades_table)

    elements.append(Spacer(1, 0.3*inch))

    if attendance:
        elements.append(Paragraph("RESUMO DE FREQUÊNCIA", ParagraphStyle(
            'SubHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        )))
        elements.append(Spacer(1, 0.2*inch))

        attendance_list = list(attendance)
        present = sum(1 for a in attendance_list if a.status == AttendanceStatus.PRESENT)
        absent = sum(1 for a in attendance_list if a.status == AttendanceStatus.ABSENT)
        justified = sum(
            1 for a in attendance_list if a.status == AttendanceStatus.EXCUSED_ABSENCE
        )
        total = len(attendance_list)
        percent = (present / total * 100) if total > 0 else 0

        freq_data = [
            ['Total de Aulas', 'Presentes', 'Ausentes', 'Justificados', 'Frequência %'],
            [str(total), str(present), str(absent), str(justified), f"{percent:.1f}%"]
        ]

        freq_table = Table(freq_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        freq_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f0fdf4')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(freq_table)

    elements.append(Spacer(1, 0.5*inch))

    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#9ca3af'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Documento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_school_history_pdf(student, *, enrollment, grades, attendance, history=None):
    """Gera o PDF do histórico escolar (ano letivo corrente) do aluno.

    Args:
        student: instância de ``Student``.
        enrollment: matrícula ativa do aluno (``Enrollment``).
        grades: iterável de ``Grade`` do aluno (com ``subject`` e ``academic_period``).
        attendance: iterável de ``Attendance`` do aluno.
        history: ``SchoolHistory`` consolidado, se houver.
    """
    from collections import OrderedDict

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
    )

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'HistHeader',
        parent=styles['Heading1'],
        fontSize=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
    )
    subheader_style = ParagraphStyle(
        'HistSubHeader',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=2,
    )
    section_style = ParagraphStyle(
        'HistSection',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1f2937'),
        spaceBefore=14,
        spaceAfter=8,
        fontName='Helvetica-Bold',
    )

    school = enrollment.school_class.school
    department = getattr(school, 'education_department', None)
    academic_year = getattr(enrollment.academic_year, 'year', None) or '—'

    elements = []
    if department is not None:
        elements.append(
            Paragraph('SECRETARIA MUNICIPAL DE EDUCAÇÃO', subheader_style)
        )
    elements.append(Paragraph(school.name, header_style))
    if school.inep_code:
        elements.append(Paragraph(f'Código INEP: {school.inep_code}', subheader_style))
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph('HISTÓRICO ESCOLAR', header_style))
    elements.append(Paragraph(f'Ano letivo {academic_year}', subheader_style))
    elements.append(Spacer(1, 0.2 * inch))

    # ------------------------------------------------------------------ aluno
    info_data = [
        ['Nome do aluno:', _student_display_name(student)],
        ['Data de nascimento:', student.birth_date.strftime('%d/%m/%Y') if student.birth_date else '—'],
        ['CPF:', format_cpf(student.cpf) if student.cpf else 'Não informado'],
        ['ID municipal:', student.unique_municipal_id],
        ['Matrícula:', enrollment.enrollment_number],
        ['Turma:', enrollment.school_class.name],
        ['Nome da mãe:', student.mother_name or '—'],
    ]
    info_table = Table(info_data, colWidths=[2 * inch, 4.6 * inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4b5563')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)

    # -------------------------------------------------------------- desempenho
    elements.append(Paragraph('Desempenho por disciplina', section_style))

    grades_list = list(grades)
    periods = OrderedDict()
    for g in sorted(
        grades_list,
        key=lambda x: getattr(x.academic_period, 'period_number', 0) or 0,
    ):
        periods.setdefault(g.academic_period_id, g.academic_period.name)
    period_ids = list(periods.keys())
    period_names = list(periods.values())

    by_subject = OrderedDict()
    for g in sorted(grades_list, key=lambda x: x.subject.name):
        by_subject.setdefault(g.subject.name, {})[g.academic_period_id] = g

    if by_subject:
        head = ['Disciplina', *period_names, 'Média']
        table_data = [head]
        for subject_name, period_map in by_subject.items():
            row = [subject_name]
            effective_values = []
            for pid in period_ids:
                grade = period_map.get(pid)
                if grade is None:
                    row.append('—')
                    continue
                eff = grade.get_effective_score()
                row.append(f'{eff}' if eff is not None else '—')
                if eff is not None:
                    effective_values.append(float(eff))
            if effective_values:
                row.append(f'{sum(effective_values) / len(effective_values):.1f}')
            else:
                row.append('—')
            table_data.append(row)

        col_count = len(head)
        first_col = 2.2 * inch
        other_col = (6.6 * inch - first_col) / (col_count - 1)
        grades_table = Table(
            table_data,
            colWidths=[first_col, *[other_col] * (col_count - 1)],
        )
        grades_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(grades_table)
    else:
        elements.append(Paragraph('Sem notas lançadas para este ano letivo.', styles['Normal']))

    # -------------------------------------------------------------- frequência
    elements.append(Paragraph('Frequência', section_style))
    attendance_list = list(attendance)
    total = len(attendance_list)
    present = sum(1 for a in attendance_list if a.status == AttendanceStatus.PRESENT)
    absent = sum(1 for a in attendance_list if a.status == AttendanceStatus.ABSENT)
    excused = sum(1 for a in attendance_list if a.status == AttendanceStatus.EXCUSED_ABSENCE)

    if history is not None and history.total_classes:
        total_classes = history.total_classes
        absences = history.absences
        percent = history.attendance_percentage
    else:
        total_classes = total
        absences = absent + excused
        percent = (present / total * 100) if total else 100.0

    freq_data = [
        ['Total de aulas', 'Faltas', 'Frequência'],
        [str(total_classes), str(absences), f'{percent:.1f}%'],
    ]
    freq_table = Table(freq_data, colWidths=[2.2 * inch, 2.2 * inch, 2.2 * inch])
    freq_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    elements.append(freq_table)

    # ------------------------------------------------------------ situação final
    status_map = {'approved': 'APROVADO', 'failed': 'REPROVADO', 'pending': 'CURSANDO'}
    final_status = status_map.get(getattr(history, 'final_status', None), 'CURSANDO')
    overall = getattr(history, 'overall_average', None)
    elements.append(Paragraph('Situação final', section_style))
    situacao_text = f'<b>{final_status}</b>'
    if overall is not None:
        situacao_text += f' &nbsp;·&nbsp; Média geral: {overall:.1f}'
    elements.append(Paragraph(situacao_text, styles['Normal']))

    # --------------------------------------------------------------- assinaturas
    elements.append(Spacer(1, 0.7 * inch))
    sign_data = [
        ['_______________________________', '_______________________________'],
        ['Diretor(a) Escolar', 'Secretário(a) Escolar'],
    ]
    sign_table = Table(sign_data, colWidths=[3.3 * inch, 3.3 * inch])
    sign_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 1), (-1, 1), 6),
    ]))
    elements.append(sign_table)

    elements.append(Spacer(1, 0.3 * inch))
    footer_style = ParagraphStyle(
        'HistFooter',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#9ca3af'),
        alignment=TA_CENTER,
    )
    elements.append(Paragraph(
        f'Documento gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")} · '
        f'Histórico do ano letivo {academic_year}',
        footer_style,
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_student_card_pdf(student):
    """Gera PDF da carteirinha do aluno com QR Code"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(4*inch, 6*inch), topMargin=0.25*inch, bottomMargin=0.25*inch)

    styles = getSampleStyleSheet()
    elements = []
    display_name = _student_display_name(student)

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(
        f"CPF:{student.cpf or ''}|MAT:{student.unique_municipal_id}|{display_name}"
    )
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)

    header_data = [['CARTEIRINHA ESCOLAR']]
    header_table = Table(header_data, colWidths=[3.5*inch])
    header_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica-Bold', 14),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#3b82f6')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.2*inch))

    info_text = f"""
    <b>Nome:</b> {display_name}<br/>
    <b>CPF:</b> {format_cpf(student.cpf) if student.cpf else '—'}<br/>
    <b>ID Municipal:</b> {student.unique_municipal_id}<br/>
    <b>Data de Nascimento:</b> {student.birth_date.strftime('%d/%m/%Y')}<br/>
    """

    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))

    from reportlab.platypus import Image as RLImage
    qr_img_rl = RLImage(qr_buffer, width=1.5*inch, height=1.5*inch)
    elements.append(qr_img_rl)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(students, grades):
    """Gera relatório em Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boletim"

        ws['A1'] = "RELATÓRIO DE NOTAS - BOLETIM CONSOLIDADO"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')

        headers = ['Nome', 'ID Municipal', 'Disciplina', 'Período', 'Nota', 'Recuperação', 'Efetiva']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row = 4
        for grade in grades:
            student = grade.enrollment.student
            ws.cell(row=row, column=1).value = _student_display_name(student)
            ws.cell(row=row, column=2).value = student.unique_municipal_id
            ws.cell(row=row, column=3).value = grade.subject.name
            ws.cell(row=row, column=4).value = grade.academic_period.name
            ws.cell(row=row, column=5).value = float(grade.score)
            ws.cell(row=row, column=6).value = (
                float(grade.recovery_score) if grade.recovery_score is not None else None
            )
            ws.cell(row=row, column=7).value = float(grade.get_effective_score())
            row += 1

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except ImportError:
        return None


def generate_csv_report(students, grades):
    """Gera relatório em CSV"""
    import csv

    buffer = io.BytesIO()
    text_buffer = io.StringIO()

    writer = csv.writer(text_buffer)
    writer.writerow([
        'Nome',
        'ID Municipal',
        'Disciplina',
        'Período',
        'Nota',
        'Recuperação',
        'Efetiva',
    ])

    for grade in grades:
        student = grade.enrollment.student
        writer.writerow([
            _student_display_name(student),
            student.unique_municipal_id,
            grade.subject.name,
            grade.academic_period.name,
            grade.score,
            grade.recovery_score if grade.recovery_score is not None else '',
            grade.get_effective_score(),
        ])

    buffer.write(text_buffer.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer
